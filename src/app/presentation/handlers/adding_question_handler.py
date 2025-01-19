from openpyxl.reader.excel import load_workbook
from telebot import TeleBot
from telebot.types import Message

from app.presentation.handlers.base import IHandler


class AddingQuestionHandler(IHandler):
    def __call__(self, message: Message, bot: TeleBot) -> None:
        bot.send_message(
            message.from_user.id,
            'Спасибо за вопрос, он добавлен в базу. '
            'В скором времени на него ответят и вам придет уведомление.'
        )
        self._add_question(message.from_user.id, message.text)
        bot.send_message(
            self.ioc.id_admin,
            'Пользователь задал вопрос'
        )

    def _add_question(self, user_id: int, question: str) -> None:
        # TODO: добавить БД
        filename = 'documents/questions.xlsx'
        wb = load_workbook(self.ioc.path / filename)
        sheet = wb.active
        count = sheet.max_row
        data = (count, user_id, False, question)
        sheet.append(data)
        wb.save(self.ioc.path / filename)
