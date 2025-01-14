from openpyxl.reader.excel import load_workbook
from telebot import TeleBot
from telebot.types import Message, InlineKeyboardMarkup, InlineKeyboardButton

from app.constants import BASE_DIR, QUESTION_TEXT
from app.utils import excel_to_2d_array


def incoming_question_handler(message: Message, bot: TeleBot) -> None:
    df = excel_to_2d_array('questions.xlsx')
    page = 1
    count = len(df) - 1

    markup = InlineKeyboardMarkup()
    markup.add(
        InlineKeyboardButton(
            text='Ответить',
            callback_data=f'{{"method":"answer", '
                          f'"NumberPage:{page}}}',
        )
    )
    markup.add(
        InlineKeyboardButton(
            text=f'{page}/{count}',
            callback_data='  '
        ),
        InlineKeyboardButton(
            text='Вперёд --->',
            callback_data=f'{{"method":"questions", '
                          f'"NumberPage":{page + 1}, '
                          f'"CountPage"{count}}}'
        )
    )

    bot.send_message(
        message.from_user.id,
        f'<b>{df[3][page]}</b>\n\n',
        parse_mode='HTML',
        reply_markup=markup,
    )


def question_handler(message: Message, bot: TeleBot) -> None:
    bot.send_message(message.from_user.id, 'Отправьте интересующий вас вопрос')
    bot.register_next_step_handler(message, add_question_handler)
    bot.send_message(message.from_user.id, QUESTION_TEXT)


def add_question_handler(message: Message) -> None:
    filename = 'documents/questions.xlsx'
    wb = load_workbook(BASE_DIR / filename)
    sheet = wb.active
    rows_count = sheet.max_row
    data = (rows_count, message.from_user.id, False, message.text)
    sheet.append(data)
    wb.save(filename=BASE_DIR / filename)
