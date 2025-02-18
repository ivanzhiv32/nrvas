from openpyxl.reader.excel import load_workbook
from telebot import TeleBot
from telebot.types import Message, ReplyKeyboardMarkup, KeyboardButton
from docxtpl import DocxTemplate

from app.domain.candidate import Candidate
from app.presentation.buttons import get_main_keyboard
from app.presentation.handlers.base import IHandler


class SendingDocumentHandler(IHandler):
    def __call__(self, message: Message, bot: TeleBot) -> None:
        try:
            phone = message.contact.phone_number
        except AttributeError:
            bot.send_message(
                message.chat.id,
                parse_mode='HTML',
                text='Для связи с Вами, нам необходимо получить Ваш номер телефона. '
                     'Нажмите кнопку в меню или напишите его в чат',
                reply_markup=self._get_keyboard()
            )
            self.next_handler(message, bot, SendingDocumentHandler(self.ioc))
            bot.delete_message(
                chat_id=message.chat.id,
                message_id=message.id - 1
            )
            return
        user_id = message.from_user.id

        usecase = self.ioc.candidate_usecase()
        with bot.retrieve_data(user_id) as data:
            candidate = Candidate(
                user_id=user_id,
                type_recruitment=data['type_recruitment'],
                nationality=data['nationality'],
                surname=data['surname'],
                name=data['name'],
                patronymic=data['patronymic'],
                birthdate=data['birthdate'],
                military_station=data['military_station'],
                university=data['university'],
                field_study=data['field_study'],
                average_score=data['average_score'],
                find_out=data['find_out'],
                phone_number=phone
            )
        doc = DocxTemplate(r'documents/Шаблон собеседования.docx')
        context = {
            'surname': candidate.surname,
            'name': candidate.name,
            'patronymic': candidate.patronymic,
            'birthday': candidate.birthdate,
            'phone_number': candidate.phone_number,
            'nationality': 'Российская Федерация',
            'university': candidate.university,
            'specialization': candidate.field_study,
            'average_score': candidate.average_score
        }
        doc.render(context)

        path = self.ioc.path
        doc.save(str(path / 'documents/Лист собеседования.docx'))

        bot.send_message(
            message.from_user.id,
            'Поздравляем, Ваша кандидатура будет рассмотрена для поступления '
            'в научную роту Военной академии связи им С.М. Буденного',
            reply_markup=get_main_keyboard(user_id == self.ioc.id_admin)
        )
        self._send_documents(message, bot)
        usecase.add_candidate(candidate)
        bot.send_message(
            self.ioc.id_admin,
            text=('<b>Зарегистрована новая заявка на поступление</b>\n'
                 f'ФИО: {candidate.surname} {candidate.name} {candidate.patronymic}\n'
                 f'Учебное заведение: {candidate.university}\n'
                 f'Номер телефона: {candidate.phone_number}'),
            parse_mode='HTML'
        )

    def _send_documents(self, message: Message, bot: TeleBot) -> None:
        chat_id = message.chat.id
        path = self.ioc.path

        with open(path / 'documents/Лист собеседования.docx', 'rb') as file:
            bot.send_document(chat_id, file)

        with open(path / 'documents/Согласие.docx', 'rb') as file:
            bot.send_document(chat_id, file)

    def _add_to_excel(self, candidate: Candidate) -> None:
        path = self.ioc.path
        filename = 'documents/candidate.xlsx'
        wb = load_workbook(path / filename)
        sheet = wb.active
        rows_count = sheet.max_row
        data = (
            rows_count,
            candidate.surname,
            candidate.name,
            candidate.patronymic,
            candidate.birthdate,
            candidate.military_station,
            candidate.university,
            candidate.field_study,
            candidate.average_score
        )
        sheet.append(data)
        wb.save(filename)

    def _get_keyboard(self) -> ReplyKeyboardMarkup:
        return ReplyKeyboardMarkup(
            resize_keyboard=True,
            row_width=3,
            one_time_keyboard=False
        ).add(
            KeyboardButton(
                'Отправить номер',
                request_contact=True
            )
        )
