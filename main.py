import json
import pandas as pd
import requests
import telebot.types
from datetime import datetime

from telebot import types
from telebot.types import ReplyKeyboardRemove
from telebot.types import ReplyKeyboardMarkup, InlineKeyboardMarkup, InlineKeyboardButton
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from time import sleep

# https://api.telegram.org/bot7541371307:AAHfbOo47fs-eSGfVN63CeCg_bUvMKRaIRI/getUpdates
TOKEN = '7541371307:AAHfbOo47fs-eSGfVN63CeCg_bUvMKRaIRI'
bot = telebot.TeleBot(TOKEN)

time = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
print(f"{time}\nProcessing...")


class Candidate:
    idUser = 0
    surname = ''
    name = ''
    patronymic = ''
    date_birth = ''
    mil_station = ''
    university = ''
    field_study = ''
    average_score = ''
    status = ''
    type_recruitment = ''
    find_out = ''
    phone_number = ''

    def add_to_excel(self):
        wb = load_workbook('documents/candidate.xlsx')
        sheet = wb.active
        rows_count = sheet.max_row
        data = (rows_count, self.surname, self.name, self.patronymic, self.date_birth,
                self.mil_station, self.university, self.field_study, self.average_score,
                self.status, self.type_recruitment, self.find_out, self.phone_number)
        sheet.append(data)

        wb.save(filename='documents/candidate.xlsx')


candidate = Candidate()


@bot.message_handler(commands=["start"])
def start(message):
    sticker = open('stickers/welcome_bender.tgs', 'rb')
    bot.send_sticker(message.chat.id, sticker)

    if message.from_user.last_name is None:
        welcome_name = f'{message.from_user.first_name}'
    else:
        welcome_name = f'{message.from_user.first_name} {message.from_user.last_name}'

    welcome_msg = f'Здравствуйте, {welcome_name}\n Я - SciComBot, и я провожу отбор кандидатов в научную роту.\nВыберете интересующий Вас раздел 👇'

    go_to_main_menu(message, welcome_msg)


def go_to_main_menu(message, msg):
    if message.from_user.id == 510572383:
        is_admin = True
    else:
        is_admin = False

    kb_markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)

    kb_about = types.KeyboardButton("#О_нас")
    kb_faq = types.KeyboardButton("FAQ")
    kb_tg_channel = types.KeyboardButton("Telegram-канал")
    kb_ask = types.KeyboardButton('Задать вопрос')
    kb_get_reg = types.KeyboardButton("Зарегистрироваться")
    kb_get_docs = types.KeyboardButton("Руководящие документы")
    kb_question = types.KeyboardButton("Входящие вопросы")

    if is_admin == True:
        kb_markup.add(kb_get_reg).row(kb_faq, kb_tg_channel, kb_about).add(kb_get_docs).add(kb_ask).add(kb_question)
    else:
        kb_markup.add(kb_get_reg).row(kb_faq, kb_tg_channel, kb_about).add(kb_get_docs).add(kb_ask)
    bot.send_message(message.chat.id, msg.format(message.from_user), parse_mode='html',
                     reply_markup=kb_markup)


@bot.message_handler(content_types=["text"])
def welcome(message):
    if message.text == 'Зарегистрироваться':
        bot.send_message(message.chat.id, 'Для прохождения регистрации заполните следующую информацию:', reply_markup=ReplyKeyboardRemove())
        type_recruitment(message)
    elif message.text == 'Telegram-канал':
        bot.send_message(message.chat.id, "https://t.me/+ntFED2PMwUo2MDZi")
    elif message.text == '#О_нас':
        bot.send_message(message.chat.id, "Данная функция находится в стадии разработки")
    elif message.text == 'Входящие вопросы':
        df = excel_to_2d_array('questions.xlsx')
        page = 1
        count = len(df) - 1

        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton(text='Ответить', callback_data="{\"method\":\"answer\",\"NumberPage\":"
                                                                       + str(page) + "}"))
        markup.add(InlineKeyboardButton(text=f'{page}/{count}', callback_data=f' '),
                   InlineKeyboardButton(text=f'Вперёд --->',
                                        callback_data="{\"method\":\"questions\",\"NumberPage\":" + str(
                                            page + 1) + ",\"CountPage\":" + str(count) + "}"))

        bot.send_message(message.from_user.id, f'<b>{df[3][page]}</b>\n\n',
                         parse_mode="HTML", reply_markup=markup)
    elif message.text == 'FAQ':
        df = excel_to_2d_array('faq.xlsx')
        # count = len(df) - 1
        page = 1

        count_pages = (len(df) - 1) // 4
        if (len(df) - 1) % 4 != 0:
            count_pages += 1

        markup = InlineKeyboardMarkup()
        i = 1
        while i <= 4:
            question = df[1][i]
            answer = df[2][i]
            markup.add(
                InlineKeyboardButton(text=question, callback_data='{\"method\":\"faq\",\"index\":' + str(i) + '}'))
            i += 1
        markup.add(InlineKeyboardButton(text='Скрыть', callback_data='unseen'))
        markup.add(InlineKeyboardButton(text=f'{page}/{count_pages}', callback_data=f' '),
                   InlineKeyboardButton(text=f'Вперёд --->',
                                        callback_data="{\"method\":\"pagination\",\"NumberPage\":" + str(
                                            page + 1) + ",\"IndexQuestion\":" + str(i) + "}"))
        bot.send_message(message.chat.id, text=f'<b>Выберите интересующий Вас вопрос</b>', parse_mode="HTML",
                         reply_markup=markup)

    elif message.text == 'Руководящие документы':
        markup = types.InlineKeyboardMarkup()
        button1 = types.InlineKeyboardButton("Ссылка", url='https://www.consultant.ru/document/cons_doc_LAW_18260/')
        markup.add(button1)

        bot.send_message(message.chat.id,
                         text='ФЗ № 53 «О воинской обязанности и военной службе» от 28.03.1998 (ред. 02.10.2024)',
                         reply_markup=markup)
    elif message.text == 'Задать вопрос':
        # kb_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        bot.send_message(message.from_user.id, 'Отправьте интересующий Вас вопрос', reply_markup=ReplyKeyboardRemove())
        bot.register_next_step_handler(message, get_question)
    elif message.text == 'id':
        candidate.idUser = message.from_user.id
        bot.send_message(message.chat.id, text=str(candidate.idUser))


def get_question(message):
    # Добавление вопроса в Excel таблицу (Номер вопроса, id пользовтеля, метка о готовности ответа, вопрос, ответ)
    id = question_to_excel(message)
    # bot.send_message(message.from_user.id,
                     # text='Спасибо за вопрос, он добавлен в базу. В скором времени на него ответят и вам придет уведомление.')
    go_to_main_menu(message, 'Спасибо за вопрос, он добавлен в базу. В скором времени на него ответят и вам придет уведомление.')

    # markup = types.InlineKeyboardButton()
    # markup.add(InlineKeyboardButton(text='Посмотреть', callback_data="{\"method\":\"pagination\",\"NumberPage\":" + str(
    #                                                     page + 1) + ",\"IndexQuestion\":" + str(index) + "}"))
    bot.send_message(510572383, 'Поступил новый вопрос от пользователя!')


def get_answer(message, page, question, id):
    # Добавление ответа к строке с вопросом в Excel таблице (Номер вопроса, id пользовтеля, метка о готовности ответа, вопрос, ответ)
    del_question(page + 1)


    bot.send_message(id, text=f'<b>Представитель научной роты ответил на ваш вопрос.</b>\n' + question + '\n' + message.text, parse_mode="HTML")
    go_to_main_menu(message, 'Ответ на данный вопрос отправлен пользователю')


# def updateVolumes(df, number, id_user, is_ready, question, answer):
#     try:
#        df.loc[author][item] += amount
#     except KeyError:
#        df = pd.concat([df,pd.DataFrame([amount], index=[author], columns=[item])]).fillna(0)

def del_question(id):
    wb = load_workbook('documents/questions.xlsx')
    sheet = wb.active
    rows_count = sheet.max_row
    sheet.delete_rows(id)

    wb.save(filename='documents/questions.xlsx')

def question_to_excel(message):
    wb = load_workbook('documents/questions.xlsx')
    sheet = wb.active
    rows_count = sheet.max_row
    data = (rows_count, message.from_user.id, False, message.text)
    sheet.append(data)

    wb.save(filename='documents/questions.xlsx')
    return rows_count


def get_surname(message):
    candidate.surname = message.text
    bot.send_message(message.from_user.id, 'Напишите ваше имя')
    bot.register_next_step_handler(message, get_name)
    bot.delete_message(chat_id=message.chat.id, message_id=message.id - 1)


def get_name(message):
    candidate.name = message.text
    bot.send_message(message.from_user.id, 'Напишите ваше отчество')
    bot.register_next_step_handler(message, get_patronymic)
    bot.delete_message(chat_id=message.chat.id, message_id=message.id - 1)


def get_patronymic(message):
    candidate.patronymic = message.text
    bot.send_message(message.from_user.id, 'Напишите название своего военного комиссариата')
    bot.register_next_step_handler(message, get_military_station)
    bot.delete_message(chat_id=message.chat.id, message_id=message.id - 1)


def get_military_station(message):
    candidate.mil_station = message.text
    bot.send_message(message.from_user.id, 'Напишите название своего ВУЗа')
    bot.register_next_step_handler(message, get_university)
    bot.delete_message(chat_id=message.chat.id, message_id=message.id - 1)


def get_university(message):
    candidate.university = message.text
    bot.send_message(message.from_user.id, 'Напишите направление подготовки в ВУЗе')
    bot.register_next_step_handler(message, get_field_study)
    bot.delete_message(chat_id=message.chat.id, message_id=message.id - 1)


def get_field_study(message):
    candidate.field_study = message.text
    bot.send_message(message.from_user.id, 'Напишите средний балл по диплому (х.х)')
    bot.register_next_step_handler(message, get_average_score)
    bot.delete_message(chat_id=message.chat.id, message_id=message.id - 1)


def get_average_score(message):
    try:
        score = float(message.text.replace(',', '.'))
        if score < 4:
            bot.send_message(message.from_user.id, 'Ваш средний балл не соответствует требованиям.\n'
                                                   'В научную роту рассматриваются кандидаты со средним баллом не менее 4.0')
            return
        elif score > 5:
            bot.send_message(message.from_user.id, 'Введен некорректный балл, попробуйте снова')
            bot.register_next_step_handler(message, get_average_score)
            return
        candidate.average_score = message.text
        bot.send_message(message.from_user.id, 'Откуда вы узнали о нас?')
        bot.register_next_step_handler(message, get_find_out)
        bot.delete_message(chat_id=message.chat.id, message_id=message.id - 1)
    except ValueError:
        bot.send_message(message.from_user.id, 'Введен некорректный балл, попробуйте снова')
        bot.register_next_step_handler(message, get_average_score)


def get_find_out(message):
    candidate.find_out = message.text

    kb_markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)

    kb_get_contact = types.KeyboardButton("Отправить номер", request_contact=True)

    kb_markup.add(kb_get_contact)
    bot.send_message(message.chat.id, parse_mode='html',
                     text='Для связи с вами, нам необходимо получить ваш номер телефона. Нажмите кнопку в меню или напишите его в чат',
                     reply_markup=kb_markup)

    bot.register_next_step_handler(message, send_docs)
    bot.delete_message(chat_id=message.chat.id, message_id=message.id - 1)


def send_docs(message):
    candidate.phone_number = message.contact.phone_number
    bot.send_message(message.from_user.id, 'Поздравляем, ваша кандидатура будет рассмотрена для поступления '
                                           'в научную роту Военной академеии связи им С.М. Буденного')

    bot.delete_message(chat_id=message.chat.id, message_id=message.id - 1)
    candidate.add_to_excel()
    go_to_main_menu(message, 'Для прохождения следующих этапов отбора заполните документы')
    bot.send_document(message.chat.id, open(r'documents/Лист собеседования.docx', 'rb'))
    bot.send_document(message.chat.id, open(r'documents/Согласие.docx', 'rb'))
    bot.send_message(510572383, text=f'<b>Зарегистрована новая заявка на поступление</b>\n' +
                                     'ФИО: ' +
                                     candidate.surname + ' ' +
                                     candidate.name + ' ' +
                                     candidate.patronymic + '\n' +
                                     'Учебное заведение: ' +
                                     candidate.university + '\n' +
                                     'Номер телефона: ' +
                                     candidate.phone_number,
                     parse_mode="HTML")


def type_recruitment(message):
    markup = types.InlineKeyboardMarkup(row_width=2)

    but_winter = types.InlineKeyboardButton("Зимний", callback_data='winter')
    but_summer = types.InlineKeyboardButton("Летний", callback_data='summer')

    markup.add(but_winter, but_summer)

    bot.send_message(message.chat.id, 'На какой призыв вы хотите оставить заявку?', reply_markup=markup)


def is_russian(message):
    markup = types.InlineKeyboardMarkup(row_width=2)

    but_yes = types.InlineKeyboardButton("Да", callback_data='yes_russian')
    but_no = types.InlineKeyboardButton("Нет", callback_data='no_russian')

    markup.add(but_yes, but_no)

    bot.edit_message_text(chat_id=message.chat.id, message_id=message.message_id,
                          text='Являетесь ли Вы гражданином Российской Федерации?', reply_markup=markup)


def is_higher_education(message):
    markup = types.InlineKeyboardMarkup(row_width=2)

    but_yes = types.InlineKeyboardButton("Да", callback_data='yes_university')
    but_no = types.InlineKeyboardButton("Нет", callback_data='no_university')

    markup.add(but_yes, but_no)

    bot.edit_message_text(chat_id=message.chat.id, message_id=message.message_id,
                          text='Имеется ли у Вас оконченное высшее техническое образование?', reply_markup=markup)


def is_aged(message):
    try:
        birthdate = datetime.strptime(message.text or "", "%d.%m.%Y")
        age = datetime.now().year - birthdate.year - (
                (datetime.now().month, datetime.now().day) < (birthdate.month, birthdate.day))
        if age < 18 or age > 30:
            go_to_main_menu(message,
                            'В Соответствии с положениями п.п."а", пункта 1, статьи 22 \"Граждане, подлежащие призыву на военную службу\"'
                            '\nФедерального закона от 28.03.1998 N 53-ФЗ (ред. от 02.10.2024) \"О воинской обязанности и военной службе\",'
                            'призыву  на  военную службу подлежат граждане мужского пола в возрасте от 18 до 30 лет, состоящие на воинском учете '
                            'или не состоящие, но обязанные состоять на воинском учете и не пребывающие в запасе ')
            return
        else:
            bot.send_message(message.from_user.id, 'Напишите вашe фамилию')
            bot.register_next_step_handler(message, get_surname)
            bot.delete_message(chat_id=message.chat.id, message_id=message.id - 1)
        candidate.date_birth = message.text
    except ValueError:
        bot.send_message(message.from_user.id, 'Введена некорректная дата, попробуйте снова')
        bot.register_next_step_handler(message, is_aged)

@bot.callback_query_handler(func=lambda call: True)
def callback_inline(call):
    req = call.data.split('_')

    try:
        if call.message:
            if call.data == 'winter':
                candidate.type_recruitment = 'Зимний'
                is_russian(call.message)
            elif call.data == 'summer':
                candidate.type_recruitment = 'Летний'
                is_russian(call.message)
            elif call.data == 'yes_russian':
                is_higher_education(call.message)
            elif call.data == 'no_russian':

                go_to_main_menu(call.message,
                                "Извините, наличие гражданства Российской Федерации "
                                "является обязательным условием для отбора в научную роту")
                bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.id)
            elif call.data == 'yes_university':
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      text='Напишите свою дату рождения в формате (дд.мм.гггг)')
                bot.register_next_step_handler(call.message, is_aged)
            elif call.data == 'no_university':
                go_to_main_menu(call.message,
                                "Извините, наличие Высшего технического образования является "
                                "обязательным условием для отбора в научную роту")
                bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.id)
            elif call.data == 'yes_age':
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      text="Заполните данные для формирования заявки на поступление")
                bot.send_message(call.message.chat.id, 'Укажите вашу фамилию')
                bot.register_next_step_handler(call.message, get_surname)
            elif call.data == 'no_age':
                go_to_main_menu(call.message,
                                'В Соответствии с положениями п.п."а", пункта 1, статьи 22 \"Граждане, подлежащие призыву на военную службу\"'
                                'Федерального закона от 28.03.1998 N 53-ФЗ (ред. от 02.10.2024) \"О воинской обязанности и военной службе\",'
                                'призыву  на  военную службу подлежат граждане мужского пола в возрасте от 18 до 30 лет, состоящие на воинском учете '
                                'или не состоящие, но обязанные состоять на воинском учете и не пребывающие в запасе ')
                bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.id)
            elif call.data == 'unseen':
                bot.delete_message(call.message.chat.id, call.message.message_id)

            elif 'pagination' in call.data:

                # Расспарсим полученный JSON
                json_string = json.loads(req[0])

                df = excel_to_2d_array('faq.xlsx')

                count_questions = (len(df) - 1)
                count_pages = (len(df) - 1) // 4
                if (len(df) - 1) % 4 != 0:
                    count_pages += 1

                page = json_string['NumberPage']

                # Пересоздаем markup
                markup = InlineKeyboardMarkup()

                index = json_string['IndexQuestion']
                finish_index = index + 3

                while index <= finish_index and index <= count_questions:
                    question = df[1][index]
                    answer = df[2][index]
                    markup.add(InlineKeyboardButton(text=question,
                                                    callback_data='{\"method\":\"faq\",\"index\":' + str(
                                                        index) + '}'))
                    index += 1

                markup.add(InlineKeyboardButton(text='Скрыть', callback_data='unseen'))

                # markup для первой страницы
                if page == 1:
                    markup.add(InlineKeyboardButton(text=f'{page}/{count_pages}', callback_data=f' '),
                               InlineKeyboardButton(text=f'Вперёд --->',
                                                    callback_data="{\"method\":\"pagination\",\"NumberPage\":" + str(
                                                        page + 1) + ",\"IndexQuestion\":" + str(index) + "}"))
                # markup для второй страницы
                elif page == count_pages:
                    markup.add(InlineKeyboardButton(text=f'<--- Назад',
                                                    callback_data="{\"method\":\"pagination\",\"NumberPage\":" + str(
                                                        page - 1) + ",\"IndexQuestion\":" + str(index - 6) + "}"),
                               InlineKeyboardButton(text=f'{page}/{count_pages}', callback_data=f' '))
                # markup для остальных страниц
                else:
                    markup.add(InlineKeyboardButton(text=f'<--- Назад',
                                                    callback_data="{\"method\":\"pagination\",\"NumberPage\":" + str(
                                                        page - 1) + ",\"IndexQuestion\":" + str(index - 8) + "}"),
                               InlineKeyboardButton(text=f'{page}/{count_pages}', callback_data=f' '),
                               InlineKeyboardButton(text=f'Вперёд --->',
                                                    callback_data="{\"method\":\"pagination\",\"NumberPage\":" + str(
                                                        page + 1) + ",\"IndexQuestion\":" + str(index) + "}"))

                # bot.send_message(call.message.chat.id, text=f'<b>Выберите интересующий вас вопрос</b>', parse_mode="HTML",
                #                  reply_markup=markup)
                bot.edit_message_text(text=f'<b>Выберите интересующий Вас вопрос:</b>', parse_mode='HTML',
                                      reply_markup=markup, chat_id=call.message.chat.id,
                                      message_id=call.message.message_id)

            elif 'faq' in call.data:
                json_string = json.loads(req[0])
                df = excel_to_2d_array('faq.xlsx')

                index = json_string['index']

                question = df[1][index]
                answer = df[2][index]

                bot.send_message(call.message.chat.id, text=f'<b>' + question + '</b>\n' + answer, parse_mode='HTML')


            elif 'answer' in call.data:
                json_string = json.loads(req[0])
                df = excel_to_2d_array('questions.xlsx')
                page = json_string['NumberPage']

                id = df[1][page]
                question = df[3][page]

                bot.send_message(call.message.chat.id, text="Отправьте ответ на вопрос:\n" + question)
                bot.register_next_step_handler(call.message, get_answer, page, question, id)

            elif 'question' in call.data:
                json_string = json.loads(req[0])

                df = excel_to_2d_array('questions.xlsx')

                page = json_string['NumberPage']
                count_pages = len(df) - 1

                question = df[3][page]

                markup = InlineKeyboardMarkup()
                markup.add(InlineKeyboardButton(text='Ответить', callback_data="{\"method\":\"answer\",\"NumberPage\":"
                                                                               + str(page) + "}"))
                # markup для первой страницы
                if page == 1:
                    markup.add(InlineKeyboardButton(text=f'{page}/{count_pages}', callback_data=f' '),
                               InlineKeyboardButton(text=f'Вперёд --->',
                                                    callback_data="{\"method\":\"question\",\"NumberPage\":" + str(
                                                        page + 1) + ",\"IndexQuestion\":" + str(page) + "}"))
                # markup для второй страницы
                elif page == count_pages:
                    markup.add(InlineKeyboardButton(text=f'<--- Назад',
                                                    callback_data="{\"method\":\"question\",\"NumberPage\":" + str(
                                                        page - 1) + ",\"IndexQuestion\":" + str(page - 1) + "}"),
                               InlineKeyboardButton(text=f'{page}/{count_pages}', callback_data=f' '))
                # markup для остальных страниц
                else:
                    markup.add(InlineKeyboardButton(text=f'<--- Назад',
                                                    callback_data="{\"method\":\"question\",\"NumberPage\":" + str(
                                                        page - 1) + ",\"IndexQuestion\":" + str(page + -1) + "}"),
                               InlineKeyboardButton(text=f'{page}/{count_pages}', callback_data=f' '),
                               InlineKeyboardButton(text=f'Вперёд --->',
                                                    callback_data="{\"method\":\"question\",\"NumberPage\":" + str(
                                                        page + 1) + ",\"IndexQuestion\":" + str(page + 1) + "}"))

                bot.edit_message_text(text=question, parse_mode='HTML',
                                      reply_markup=markup, chat_id=call.message.chat.id,
                                      message_id=call.message.message_id)
            # elif call.data == 'new_question':
            #     df = excel_to_2d_array('questions.xlsx')
            #
            #     question = df[][]
            elif call.data == 'send_phone_number':
                bot.register_next_step_handler(call.message, send_docs)

    except Exception as e:
        print(repr(e))


def excel_to_2d_array(name_doc):
    wb = load_workbook('documents/' + name_doc)
    sheet = wb.active
    df = pd.DataFrame(sheet.values)
    return df


bot.polling()

